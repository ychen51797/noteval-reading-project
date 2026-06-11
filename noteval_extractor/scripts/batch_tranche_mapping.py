"""
batch_tranche_mapping.py — Batch tranche mapping from agent export XML → Excel.

**Primary input:** ``noteval_export`` XML files (``{deal_id}_{YYYYMMDD}.xml``)
produced after extraction, e.g. via::

  py -3 noteval_extractor/scripts/export_noteval_xml.py <output_folder> --map-tranches

Input options (pick one):

1. **CSV** with ``xml_file`` (path to each XML), or ``deal_id`` + ``payment_date``
   (script resolves ``{xml_root}/{deal_id}_{YYYYMMDD}.xml``).

2. **``--xml``** one or more ``.xml`` paths.

3. **``--xml-dir``** — all ``*.xml`` in a folder (e.g. 10 deals for a pilot).

4. **``--deal`` / ``--date``** pairs (resolve under ``--xml-root``).

Usage (10-deal pilot)::

  py -3 noteval_extractor/scripts/export_noteval_xml.py output/DEAL_DATE --map-tranches
  # … repeat per deal, XML lands in noteval_extractor/xml/

  py -3 noteval_extractor/scripts/batch_tranche_mapping.py pilot_deals.csv \\
      -o tranche_mapping_10.xlsx

  py -3 noteval_extractor/scripts/batch_tranche_mapping.py --xml-dir noteval_extractor/xml \\
      -o tranche_mapping_all.xlsx --compare-db

Environment: NOTEVAL_ODBC_CONNECTION, or ``--tranche-cache`` + ``--no-db``.
``--compare-db`` (default when DB is enabled) joins ``cdo_noteval_tranches`` and
``cdo_tranche_master`` (``orig_balance`` vs export ``original_balance``, keyed by
``deal_id`` + ``tranche_id`` / ``moodystrancheid``) and writes DB values + diffs
on the Tranche mapping sheet and a Match rates summary.
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from export_noteval_xml import parse_us_date  # noqa: E402
from map_tranches import (  # noqa: E402
    DB_COMPARE_FIELDS,
    DB_COMPARE_HEADERS,
    MASTER_COMPARE_FIELDS,
    MASTER_COMPARE_HEADERS,
    TrancheMapper,
    compare_tranche_to_db,
    compare_tranche_to_master,
)

COMPARE_FIELDS = DB_COMPARE_FIELDS + MASTER_COMPARE_FIELDS

_DEAL_XML_STEM = re.compile(r"^(\d+)_(\d{8})(?:_sdk|_llm)?$")

_MAPPING_HEADERS = [
    "deal_id",
    "payment_date",
    "xml_file",
    "schema_version",
    "deal_name",
    "class_name",
    "map_class",
    "cusip",
    "moodystrancheid",
    "trustee_tranche_name",
    "map_tier",
    "map_status",
    "map_message",
    "xml_map_status",
    "interest_rate",
    "original_balance",
    "beginning_balance",
    "interest_payment",
    "principal_payment",
    "deferred_interest",
    "ending_balance",
]

_SUMMARY_HEADERS = [
    "deal_id",
    "payment_date",
    "xml_file",
    "schema_version",
    "status",
    "class_rows",
    "mapped_ok",
    "unmapped",
    "ambiguous",
    "already_mapped_in_xml",
    "notes",
]

_MATCH_RATE_HEADERS = [
    "deal_id",
    "payment_date",
    "xml_file",
    "class_rows",
    "tranches_mapped",
    "tranche_map_rate_pct",
    "interest_rate_compared",
    "interest_rate_match_pct",
    "interest_payment_compared",
    "interest_payment_match_pct",
    "principal_payment_compared",
    "principal_payment_match_pct",
    "deferred_interest_compared",
    "deferred_interest_match_pct",
    "beginning_balance_compared",
    "beginning_balance_match_pct",
    "ending_balance_compared",
    "ending_balance_match_pct",
    "original_balance_compared",
    "original_balance_match_pct",
    "notes",
]


@dataclass
class DealMatchStats:
    deal_id: str
    payment_date: str
    xml_file: str
    class_rows: int = 0
    mapped_ok: int = 0
    field_match: dict[str, int] = field(default_factory=lambda: {f: 0 for f in COMPARE_FIELDS})
    field_compared: dict[str, int] = field(default_factory=lambda: {f: 0 for f in COMPARE_FIELDS})

    def tranche_map_rate_pct(self) -> float | None:
        if self.class_rows <= 0:
            return None
        return 100.0 * self.mapped_ok / self.class_rows

    def field_match_rate_pct(self, field: str) -> float | None:
        n = self.field_compared.get(field, 0)
        if n <= 0:
            return None
        return 100.0 * self.field_match.get(field, 0) / n


def _pct_cell(value: float | None) -> Any:
    return round(value, 1) if value is not None else ""


def _record_db_diff(stats: DealMatchStats, diffs: dict[str, str]) -> None:
    for field in DB_COMPARE_FIELDS:
        diff = str(diffs.get(f"diff_{field}") or "").strip()
        if diff == "match":
            stats.field_match[field] += 1
            stats.field_compared[field] += 1
        elif diff:
            stats.field_compared[field] += 1


def _record_master_diff(stats: DealMatchStats, diffs: dict[str, str]) -> None:
    diff = str(diffs.get("diff_original_balance") or "").strip()
    if diff == "match":
        stats.field_match["original_balance"] += 1
        stats.field_compared["original_balance"] += 1
    elif diff:
        stats.field_compared["original_balance"] += 1


def _match_stats_to_row(stats: DealMatchStats, notes: str = "") -> list[Any]:
    row: list[Any] = [
        stats.deal_id,
        stats.payment_date,
        stats.xml_file,
        stats.class_rows,
        stats.mapped_ok,
        _pct_cell(stats.tranche_map_rate_pct()),
    ]
    for field in COMPARE_FIELDS:
        row.append(stats.field_compared.get(field, 0))
        row.append(_pct_cell(stats.field_match_rate_pct(field)))
    row.append(notes)
    return row


def _aggregate_match_stats(stats_list: list[DealMatchStats]) -> DealMatchStats:
    total = DealMatchStats(deal_id="TOTAL", payment_date="", xml_file="")
    for s in stats_list:
        total.class_rows += s.class_rows
        total.mapped_ok += s.mapped_ok
        for f in COMPARE_FIELDS:
            total.field_match[f] += s.field_match[f]
            total.field_compared[f] += s.field_compared[f]
    return total


# Excel presentation (Compare-with-DB workbook)
_FILL_HEADER = "F0F4F8"
_FILL_OK = "E8F5E9"
_FILL_WARN = "FFF8E1"
_FILL_BAD = "FFEBEE"
_FILL_MATCH = "C8E6C9"
_FILL_DIFF = "FFCDD2"
_FILL_DB_HDR = "E3F2FD"
_HIDDEN_MAPPING_COLS = frozenset({"xml_file", "schema_version", "map_message"})
_HIDDEN_MATCH_RATE_COLS = frozenset({"xml_file"})
_DIFF_HEADERS = {f"diff_{f}" for f in DB_COMPARE_FIELDS} | {"diff_original_balance"}
_DB_HEADERS = {f"db_{f}" for f in DB_COMPARE_FIELDS} | {"db_orig_balance"}
_PCT_SUFFIX = "_match_pct"
_MAP_STATUS_OK = frozenset({"ok"})
_MAP_STATUS_BAD = frozenset({"unmapped", "no_data", "missing"})
_MAP_STATUS_WARN = frozenset({"ambiguous", "partial"})


def _header_index(headers: list[str]) -> dict[str, int]:
    return {h: i for i, h in enumerate(headers)}


def _row_has_field_mismatch(row: list[Any], headers: list[str]) -> bool:
    idx = _header_index(headers)
    for h in _DIFF_HEADERS:
        i = idx.get(h)
        if i is None or i >= len(row):
            continue
        v = str(row[i] or "").strip()
        if v and v != "match":
            return True
    return False


def _mapping_mismatch_rows(headers: list[str], rows: list[list[Any]]) -> list[list[Any]]:
    return [r for r in rows if _row_has_field_mismatch(r, headers)]


def _autosize_columns(ws, headers: list[str], rows: list[list[Any]], *, cap: int = 48) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in rows[:500]:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1] or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), cap)


def _style_compare_sheet(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    sheet_kind: str,
) -> None:
    """Apply headers, filters, colors, and column hiding for readability."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill_header = PatternFill("solid", fgColor=_FILL_HEADER)
    fill_db_hdr = PatternFill("solid", fgColor=_FILL_DB_HDR)
    fill_ok = PatternFill("solid", fgColor=_FILL_OK)
    fill_warn = PatternFill("solid", fgColor=_FILL_WARN)
    fill_bad = PatternFill("solid", fgColor=_FILL_BAD)
    fill_match = PatternFill("solid", fgColor=_FILL_MATCH)
    fill_diff = PatternFill("solid", fgColor=_FILL_DIFF)

    idx = _header_index(headers)
    hide_cols = _HIDDEN_MAPPING_COLS if sheet_kind in ("mapping", "mismatches") else _HIDDEN_MATCH_RATE_COLS

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = fill_db_hdr if header in _DB_HEADERS or header in _DIFF_HEADERS else fill_header
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if header in hide_cols:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    map_status_i = idx.get("map_status")
    db_status_i = idx.get("db_match_status")
    tm_status_i = idx.get("tm_match_status")
    diff_is = [idx[h] for h in _DIFF_HEADERS if h in idx]
    pct_is = [idx[h] for h in headers if h.endswith(_PCT_SUFFIX)]
    tranche_map_i = idx.get("tranche_map_rate_pct")

    for row_num in range(2, ws.max_row + 1):
        deal_cell = ws.cell(row=row_num, column=1)
        if str(deal_cell.value or "").strip().upper() == "TOTAL":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).font = Font(bold=True)
            continue

        if map_status_i is not None:
            cell = ws.cell(row=row_num, column=map_status_i + 1)
            st = str(cell.value or "").strip().lower()
            if st in _MAP_STATUS_OK:
                cell.fill = fill_ok
            elif st in _MAP_STATUS_BAD:
                cell.fill = fill_bad
            elif st in _MAP_STATUS_WARN:
                cell.fill = fill_warn

        if db_status_i is not None:
            cell = ws.cell(row=row_num, column=db_status_i + 1)
            st = str(cell.value or "").strip().lower()
            if st == "ok":
                cell.fill = fill_ok
            elif st in ("no_match", "no_data"):
                cell.fill = fill_bad

        if tm_status_i is not None:
            cell = ws.cell(row=row_num, column=tm_status_i + 1)
            st = str(cell.value or "").strip().lower()
            if st == "ok":
                cell.fill = fill_ok
            elif st in ("no_match", "no_data"):
                cell.fill = fill_bad

        for di in diff_is:
            cell = ws.cell(row=row_num, column=di + 1)
            v = str(cell.value or "").strip()
            if v == "match":
                cell.fill = fill_match
            elif v:
                cell.fill = fill_diff
                cell.font = Font(bold=True)

        for pi in pct_is:
            cell = ws.cell(row=row_num, column=pi + 1)
            raw = cell.value
            if raw in (None, ""):
                continue
            try:
                pct = float(raw)
            except (TypeError, ValueError):
                continue
            if pct >= 99.9:
                cell.fill = fill_ok
            elif pct >= 80:
                cell.fill = fill_warn
            else:
                cell.fill = fill_bad

        if tranche_map_i is not None:
            cell = ws.cell(row=row_num, column=tranche_map_i + 1)
            raw = cell.value
            if raw in (None, ""):
                continue
            try:
                pct = float(raw)
            except (TypeError, ValueError):
                continue
            if pct >= 99.9:
                cell.fill = fill_ok
            elif pct >= 80:
                cell.fill = fill_warn
            else:
                cell.fill = fill_bad

    _autosize_columns(ws, headers, rows)


def _write_compare_sheet(wb, title: str, headers: list[str], rows: list[list[Any]], *, sheet_kind: str, index: int):
    ws = wb.create_sheet(title, index)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_compare_sheet(ws, headers, rows, sheet_kind=sheet_kind)
    return ws


@dataclass
class XmlBatchItem:
    xml_path: Path
    deal_id: str
    payment_date_input: str = ""
    payment_iso: str = ""


def _default_xml_root() -> Path:
    return _SCRIPT_DIR.parent / "xml"


def openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


def normalize_payment_date_yyyymmdd(raw: str) -> tuple[str, str]:
    """Return (YYYYMMDD, ISO YYYY-MM-DD or '')."""
    t = (raw or "").strip()
    if not t:
        return "", ""
    if len(t) == 8 and t.isdigit():
        ymd = t
    else:
        d = parse_us_date(t)
        if not d:
            for fmt in ("%B %d, %Y", "%b %d, %Y"):
                try:
                    d = datetime.strptime(t, fmt)
                    break
                except ValueError:
                    continue
        if not d:
            digits = "".join(c for c in t if c.isdigit())
            if len(digits) == 8:
                ymd = digits
            else:
                return "", t
        else:
            ymd = f"{d.year:04d}{d.month:02d}{d.day:02d}"
    try:
        iso = datetime.strptime(ymd, "%Y%m%d").strftime("%Y-%m-%d")
    except ValueError:
        iso = ""
    return ymd, iso


def payment_iso_from_item_meta(meta: dict[str, str], ymd_file: str = "") -> str:
    """Best-effort ISO payment date for DB joins."""
    if ymd_file:
        _, iso = normalize_payment_date_yyyymmdd(ymd_file)
        if iso:
            return iso
    raw = (meta.get("payment_date") or "").strip()
    _, iso = normalize_payment_date_yyyymmdd(raw)
    return iso or raw


def parse_deal_date_from_xml_path(path: Path) -> tuple[str, str]:
    """Infer (deal_id, yyyymmdd) from ``869358159_20260420.xml`` stem."""
    stem = path.stem
    m = _DEAL_XML_STEM.match(stem)
    if m:
        return m.group(1), m.group(2)
    return "", ""


def resolve_xml_path(
    xml_root: Path,
    deal_id: str,
    payment_date_input: str,
) -> Path | None:
    pay_ymd, _ = normalize_payment_date_yyyymmdd(payment_date_input)
    if not pay_ymd:
        return None
    deal_id = deal_id.strip()
    for stem in (f"{deal_id}_{pay_ymd}", f"{deal_id}_{pay_ymd}_sdk"):
        p = xml_root / f"{stem}.xml"
        if p.is_file():
            return p.resolve()
    return None


def read_export_xml_meta(xml_path: Path) -> dict[str, str]:
    """Read deal_id, dates, deal_name, schema from noteval_export XML."""
    root = ET.parse(xml_path).getroot()
    if root.tag != "noteval_export":
        raise ValueError(f"Not a noteval_export file: {xml_path}")
    meta: dict[str, str] = {
        "schema_version": root.get("schema_version", "") or "",
        "source_folder": root.get("source_folder", "") or "",
    }
    md = root.find("metadata")
    if md is not None:
        for tag in ("deal_id", "deal_name", "payment_date", "currency"):
            el = md.find(tag)
            if el is not None and (el.text or "").strip():
                meta[tag] = el.text.strip()
            if tag == "payment_date" and el is not None:
                disp = el.get("display")
                if disp:
                    meta["payment_date_display"] = disp
    did_file, ymd_file = parse_deal_date_from_xml_path(xml_path)
    if not meta.get("deal_id") and did_file:
        meta["deal_id"] = did_file
    if ymd_file:
        meta["payment_date_ymd"] = ymd_file
    return meta


def class_lines_from_xml(xml_path: Path) -> list[dict[str, str]]:
    root = ET.parse(xml_path).getroot()
    out: list[dict[str, str]] = []
    classes = root.find("classes")
    if classes is None:
        return out
    for ce in classes.findall("class"):
        cname = (ce.get("name") or "").strip()
        if not cname:
            continue
        xml_tid = (ce.get("moodystrancheid") or "").strip()
        xml_map_status = (ce.get("map_status") or "").strip()
        xml_map_tier = (ce.get("map_tier") or "").strip()
        base = {
            "class_name": cname,
            "map_class": (ce.get("map_class") or "").strip(),
            "cusip": "",
            "xml_moodystrancheid": xml_tid,
            "xml_map_status": xml_map_status,
            "xml_map_tier": xml_map_tier,
            "interest_rate": (ce.findtext("interest_rate") or "").strip(),
            "original_balance": (ce.findtext("original_balance") or "").strip(),
            "beginning_balance": (ce.findtext("beginning_balance") or "").strip(),
            "interest_payment": (ce.findtext("interest_payment") or "").strip(),
            "principal_payment": (ce.findtext("principal_payment") or "").strip(),
            "deferred_interest": (ce.findtext("deferred_interest") or "").strip(),
            "ending_balance": (ce.findtext("ending_balance") or "").strip(),
        }
        idel = ce.find("identifiers")
        cusips: list[str] = []
        if idel is not None:
            for line in idel.findall("line"):
                c = (line.get("cusip") or "").strip()
                if c and c.upper() != "N/A":
                    cusips.append(c)
                if not xml_tid:
                    xml_tid = (line.get("moodystrancheid") or "").strip()
                    if xml_tid:
                        base["xml_moodystrancheid"] = xml_tid
        base["cusips"] = cusips
        base["cusip"] = cusips[0] if cusips else ""
        out.append(base)
    return out


def load_requests_csv(path: Path, xml_root: Path) -> list[XmlBatchItem]:
    items: list[XmlBatchItem] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return items
        fields = {h.strip().lower().replace(" ", "_"): h for h in reader.fieldnames if h}
        xml_col = fields.get("xml_file") or fields.get("xml_path") or fields.get("xml")
        did_col = fields.get("deal_id") or fields.get("moodysdealid")
        pd_col = fields.get("payment_date") or fields.get("date")

        for row in reader:
            xml_path: Path | None = None
            deal_id = ""
            pay_in = ""

            if xml_col:
                raw = (row.get(xml_col) or "").strip()
                if raw:
                    xml_path = Path(raw)
                    if not xml_path.is_absolute():
                        xml_path = (path.parent / xml_path).resolve()
            if did_col:
                deal_id = (row.get(did_col) or "").strip()
            if pd_col:
                pay_in = (row.get(pd_col) or "").strip()

            if xml_path is None and deal_id and pay_in:
                xml_path = resolve_xml_path(xml_root, deal_id, pay_in)

            if xml_path is None or not xml_path.is_file():
                raise FileNotFoundError(
                    f"XML not found for row deal_id={deal_id!r} payment_date={pay_in!r} "
                    f"(set xml_file column or place file under {xml_root})"
                )

            meta = read_export_xml_meta(xml_path)
            did = meta.get("deal_id") or deal_id or parse_deal_date_from_xml_path(xml_path)[0]
            ymd = meta.get("payment_date_ymd", "")
            if not pay_in and ymd:
                pay_in = ymd
            pay_iso = payment_iso_from_item_meta(meta, ymd)

            items.append(
                XmlBatchItem(
                    xml_path=xml_path,
                    deal_id=did,
                    payment_date_input=pay_in or pay_iso,
                    payment_iso=pay_iso,
                )
            )
    return items


def items_from_xml_paths(paths: list[Path]) -> list[XmlBatchItem]:
    items: list[XmlBatchItem] = []
    for p in paths:
        p = p.resolve()
        if not p.is_file():
            raise FileNotFoundError(p)
        meta = read_export_xml_meta(p)
        did = meta.get("deal_id") or parse_deal_date_from_xml_path(p)[0]
        ymd = meta.get("payment_date_ymd", "")
        pay_iso = payment_iso_from_item_meta(meta, ymd)
        items.append(
            XmlBatchItem(
                xml_path=p,
                deal_id=did,
                payment_date_input=ymd or pay_iso,
                payment_iso=pay_iso,
            )
        )
    return items


def items_from_xml_dir(xml_dir: Path, *, max_deals: int = 0) -> list[XmlBatchItem]:
    paths = sorted(xml_dir.glob("*.xml"), key=lambda x: x.name.lower())
    if max_deals > 0:
        paths = paths[:max_deals]
    return items_from_xml_paths(paths)


def process_xml_item(
    item: XmlBatchItem,
    mapper: TrancheMapper,
    *,
    compare_db: bool = False,
) -> tuple[list[list[Any]], list[Any], DealMatchStats | None]:
    xml_path = item.xml_path
    deal_id = item.deal_id
    notes: list[str] = []

    try:
        meta = read_export_xml_meta(xml_path)
    except (ET.ParseError, ValueError) as e:
        notes.append(str(e))
        meta = {}

    schema = meta.get("schema_version", "")
    deal_name = meta.get("deal_name", "")
    pay_iso = item.payment_iso or meta.get("payment_date", "")
    if schema and schema < "3":
        notes.append(f"old schema_version={schema}")

    class_lines = class_lines_from_xml(xml_path)
    peer_class_names = [
        str(line.get("class_name") or "").strip()
        for line in class_lines
        if str(line.get("class_name") or "").strip()
    ]
    mapping_rows: list[list[Any]] = []
    ok = unmapped = ambiguous = already = 0
    match_stats = (
        DealMatchStats(
            deal_id=deal_id,
            payment_date=pay_iso,
            xml_file=str(xml_path),
        )
        if compare_db
        else None
    )

    for line in class_lines:
        xml_tid = line.get("xml_moodystrancheid", "")
        xml_st = line.get("xml_map_status", "")

        cusips_raw = line.get("cusips") or []
        primary_cusip = (line.get("cusip") or "").strip()
        has_cusip = bool(cusips_raw) or bool(
            primary_cusip and primary_cusip.upper() != "N/A"
        )

        # Tier 1 (CUSIP) always runs when identifiers are present — do not trust
        # a prior schema-4 XML name map when listing CUSIPs exist.
        if xml_tid and xml_st == "ok" and not has_cusip:
            r_tid = xml_tid
            r_tier = line.get("xml_map_tier") or "xml"
            r_status = "ok"
            r_tname = ""
            r_msg = "from export XML (no CUSIP on class)"
            already += 1
        else:
            if cusips_raw:
                r = mapper.resolve(
                    deal_id,
                    cusips=cusips_raw,
                    class_name=line.get("class_name") or None,
                    map_class=line.get("map_class") or None,
                    peer_class_names=peer_class_names,
                    deal_name=deal_name or None,
                )
            else:
                r = mapper.resolve(
                    deal_id,
                    cusip=primary_cusip or None,
                    class_name=line.get("class_name") or None,
                    map_class=line.get("map_class") or None,
                    peer_class_names=peer_class_names,
                    deal_name=deal_name or None,
                )
            r_tid = r.moodystrancheid or ""
            r_tier = r.map_tier or ""
            r_status = r.map_status or ""
            r_tname = r.trustee_tranche_name or ""
            r_msg = r.map_message or ""
            if (
                has_cusip
                and xml_tid
                and xml_st == "ok"
                and r_tid == xml_tid
                and r_tier == "cusip"
            ):
                already += 1

        if r_status == "ok":
            ok += 1
        elif r_status == "ambiguous":
            ambiguous += 1
        else:
            unmapped += 1

        if match_stats is not None:
            match_stats.class_rows += 1
            if r_status == "ok":
                match_stats.mapped_ok += 1

        row = [
            deal_id,
            pay_iso,
            str(xml_path),
            schema,
            deal_name,
            line.get("class_name", ""),
            line.get("map_class", ""),
            line.get("cusip", ""),
            r_tid,
            r_tname,
            r_tier,
            r_status,
            r_msg,
            xml_st or ("ok" if xml_tid else ""),
            line.get("interest_rate", ""),
            line.get("original_balance", ""),
            line.get("beginning_balance", ""),
            line.get("interest_payment", ""),
            line.get("principal_payment", ""),
            line.get("deferred_interest", ""),
            line.get("ending_balance", ""),
        ]

        if compare_db:
            extracted = {
                "interest_rate": line.get("interest_rate", ""),
                "interest_payment": line.get("interest_payment", ""),
                "principal_payment": line.get("principal_payment", ""),
                "deferred_interest": line.get("deferred_interest", ""),
                "beginning_balance": line.get("beginning_balance", ""),
                "ending_balance": line.get("ending_balance", ""),
                "original_balance": line.get("original_balance", ""),
            }
            db_row, matched_name = mapper.db_row_for(
                deal_id,
                pay_iso,
                moodystrancheid=r_tid or None,
            )
            diffs = compare_tranche_to_db(extracted, db_row)
            db_status = (
                "ok"
                if db_row is not None
                else ("no_data" if not pay_iso else "no_match")
            )
            for field in DB_COMPARE_FIELDS:
                row.append(diffs.get(f"db_{field}", ""))
            row.append(matched_name)
            row.append(db_status)
            for field in DB_COMPARE_FIELDS:
                row.append(diffs.get(f"diff_{field}", ""))
            if match_stats is not None:
                _record_db_diff(match_stats, diffs)

            master_row, _master_tid = mapper.master_row_for(
                deal_id,
                moodystrancheid=r_tid or None,
            )
            mdiffs = compare_tranche_to_master(extracted, master_row)
            tm_status = (
                "ok"
                if master_row is not None
                else ("no_data" if not r_tid else "no_match")
            )
            row.append(mdiffs.get("db_orig_balance", ""))
            row.append(tm_status)
            row.append(mdiffs.get("diff_original_balance", ""))
            if match_stats is not None:
                _record_master_diff(match_stats, mdiffs)

        mapping_rows.append(row)

    if not class_lines:
        notes.append("no <class> rows in XML")
        empty_row = [
            deal_id,
            pay_iso,
            str(xml_path),
            schema,
            deal_name,
            "",
            "",
            "",
            "",
            "",
            "",
            "no_data",
            "; ".join(notes),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        if compare_db:
            empty_row.extend([""] * (len(DB_COMPARE_HEADERS) + len(MASTER_COMPARE_HEADERS)))
        mapping_rows.append(empty_row)

    if not class_lines:
        status = "missing"
    elif ambiguous or unmapped:
        status = "partial"
    else:
        status = "ok"

    summary = [
        deal_id,
        pay_iso,
        str(xml_path),
        schema,
        status,
        len(class_lines),
        ok,
        unmapped,
        ambiguous,
        already,
        "; ".join(notes),
    ]
    return mapping_rows, summary, match_stats


def _mapping_headers(*, compare_db: bool) -> list[str]:
    headers = list(_MAPPING_HEADERS)
    if compare_db:
        headers.extend(DB_COMPARE_HEADERS)
        headers.extend(MASTER_COMPARE_HEADERS)
    return headers


def build_workbook(
    items: list[XmlBatchItem],
    mapper: TrancheMapper,
    *,
    compare_db: bool = False,
) -> bytes:
    if not openpyxl_available():
        raise SystemExit(
            "openpyxl is required. Install with: py -3 -m pip install openpyxl"
        )
    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    all_mapping: list[list[Any]] = []
    summaries: list[list[Any]] = []
    match_stats_list: list[DealMatchStats] = []

    for did in sorted({x.deal_id for x in items if x.deal_id}):
        try:
            mapper.maps_for_deal(did)
        except Exception as e:
            print(f"warn: preload maps {did}: {e}", file=sys.stderr)
        if compare_db:
            try:
                mapper.tranche_master_for(did)
            except Exception as e:
                print(f"warn: preload tranche master {did}: {e}", file=sys.stderr)

    for item in items:
        mrows, summ, mstats = process_xml_item(item, mapper, compare_db=compare_db)
        all_mapping.extend(mrows)
        summaries.append(summ)
        if mstats is not None:
            match_stats_list.append(mstats)

    sheet_idx = 0
    if compare_db and match_stats_list:
        match_rows = [_match_stats_to_row(s) for s in match_stats_list]
        if len(match_stats_list) > 1:
            match_rows.append(_match_stats_to_row(_aggregate_match_stats(match_stats_list)))
        _write_compare_sheet(
            wb, "Match rates", _MATCH_RATE_HEADERS, match_rows, sheet_kind="match_rates", index=sheet_idx
        )
        sheet_idx += 1

    mapping_headers = _mapping_headers(compare_db=compare_db)
    if compare_db:
        mismatch_rows = _mapping_mismatch_rows(mapping_headers, all_mapping)
        if mismatch_rows:
            _write_compare_sheet(
                wb,
                "Mismatches",
                mapping_headers,
                mismatch_rows,
                sheet_kind="mismatches",
                index=sheet_idx,
            )
            sheet_idx += 1

    _write_compare_sheet(
        wb,
        "Tranche mapping",
        mapping_headers,
        all_mapping,
        sheet_kind="mapping",
        index=sheet_idx,
    )
    sheet_idx += 1
    _write_compare_sheet(wb, "Summary", _SUMMARY_HEADERS, summaries, sheet_kind="summary", index=sheet_idx)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Batch tranche mapping from agent noteval_export XML files."
    )
    ap.add_argument(
        "csv",
        nargs="?",
        type=Path,
        help="CSV: xml_file column OR deal_id + payment_date",
    )
    ap.add_argument(
        "--xml",
        nargs="*",
        type=Path,
        default=[],
        help="One or more export XML paths",
    )
    ap.add_argument(
        "--xml-dir",
        type=Path,
        help="Directory of *.xml export files (e.g. noteval_extractor/xml)",
    )
    ap.add_argument(
        "--xml-root",
        type=Path,
        default=None,
        help=f"Resolve deal_id+date to XML here (default: {_default_xml_root()})",
    )
    ap.add_argument(
        "--deal",
        action="append",
        default=[],
        help="Deal id (with --date; XML under --xml-root)",
    )
    ap.add_argument("--date", action="append", default=[])
    ap.add_argument(
        "--max-deals",
        type=int,
        default=0,
        help="Cap deals when using --xml-dir (0 = no cap)",
    )
    ap.add_argument("-o", "--output", type=Path, required=True, help="Output .xlsx")
    ap.add_argument("--tranche-cache", type=Path, help="JSON from map_tranches --prefetch")
    ap.add_argument("--no-db", action="store_true", help="Use only --tranche-cache")
    ap.add_argument(
        "--compare-db",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Compare XML tranche fields to cdo_noteval_tranches (default: on when DB enabled)",
    )
    args = ap.parse_args()

    xml_root = (args.xml_root or _default_xml_root()).resolve()
    items: list[XmlBatchItem] = []

    try:
        if args.csv:
            items.extend(load_requests_csv(args.csv.resolve(), xml_root))
        if args.xml:
            items.extend(items_from_xml_paths([p.resolve() for p in args.xml]))
        if args.xml_dir:
            items.extend(
                items_from_xml_dir(
                    args.xml_dir.resolve(),
                    max_deals=args.max_deals or 0,
                )
            )
        if args.deal:
            if len(args.deal) != len(args.date):
                print("--deal and --date must be paired equally", file=sys.stderr)
                return 2
            for did, pd in zip(args.deal, args.date):
                xp = resolve_xml_path(xml_root, did, pd)
                if xp is None:
                    print(f"XML not found: {did} {pd} under {xml_root}", file=sys.stderr)
                    return 1
                ymd, iso = normalize_payment_date_yyyymmdd(pd)
                items.append(
                    XmlBatchItem(
                        xml_path=xp,
                        deal_id=did.strip(),
                        payment_date_input=pd,
                        payment_iso=iso,
                    )
                )
    except FileNotFoundError as e:
        print(e, file=sys.stderr)
        return 1
    except ValueError as e:
        print(e, file=sys.stderr)
        return 1

    if not items:
        print(
            "Provide csv, --xml, --xml-dir, or --deal/--date. "
            "Input must be noteval_export XML from the agent pipeline.",
            file=sys.stderr,
        )
        return 2

    # dedupe by xml path
    seen: set[str] = set()
    unique: list[XmlBatchItem] = []
    for it in items:
        key = str(it.xml_path)
        if key not in seen:
            seen.add(key)
            unique.append(it)
    items = unique

    mapper = TrancheMapper(
        cache_file=args.tranche_cache.resolve() if args.tranche_cache else None,
        use_db=not args.no_db,
    )
    compare_db = bool(args.compare_db and not args.no_db)

    out = args.output.resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(build_workbook(items, mapper, compare_db=compare_db))
    print(out)
    mode = "mapping + DB compare" if compare_db else "mapping only"
    print(f"Deals: {len(items)}  ({mode}) — see Summary sheet", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
