"""
export_noteval_excel.py — Roll up noteval markdown outputs into one .xlsx workbook.

Sheets (DB-compare layout):

- **Classes** — deal_id, payment_date, class_name, cusip, interest_payment,
  principal_payment, beginning_balance, deferred_interest, ending_balance
  (optional: moodystrancheid, map_tier, map_status via ``--map-tranches``)
- **Valuation fees** — deal_id, payment_date, main_category, sub_category, amount_paid
- **Summary** — per-folder file / validation flags

Usage (repo root)::

  py -3 noteval_extractor/scripts/export_noteval_excel.py \\
      noteval_extractor/output/DEAL1_DATE noteval_extractor/output/DEAL2_DATE \\
      -o noteval_export.xlsx

Requires: ``pip install openpyxl``
"""

from __future__ import annotations

import argparse
import io
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit(
        "openpyxl is required for Excel export. Install with: py -3 -m pip install openpyxl"
    ) from e

_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from export_noteval_xml import (  # noqa: E402
    _col_exact,
    _norm_header,
    _parse_tables_kv,
    fee_amount_col,
    fee_data_rows,
    fee_main_col,
    fee_priority_col,
    fee_type_col,
    find_class_balance_table,
    find_listing_table,
    find_valuation_fees_table,
    parse_md_tables,
)
from map_tranches import CLASS_MAP_EXTRA_HEADERS, TrancheMapper  # noqa: E402

_DEAL_FOLDER = re.compile(r"^\d+_\d{8}(_sdk|_llm)?$")
_SUMMARY_HEADERS = [
    "folder",
    "deal_id",
    "payment_date",
    "pipeline",
    "has_01",
    "has_02",
    "has_03",
    "has_04",
    "has_05",
    "validation_status",
    "validation_errors",
    "validation_warnings",
    "deal_name",
    "payment_date_01",
    "currency",
    "class_row_count",
    "fee_row_count",
    "export_notes",
]
_CLASS_EXPORT_HEADERS = [
    "deal_id",
    "payment_date",
    "class_name",
    "cusip",
    "interest_payment",
    "principal_payment",
    "beginning_balance",
    "deferred_interest",
    "ending_balance",
]
_VALUATION_FEE_HEADERS = [
    "deal_id",
    "payment_date",
    "main_category",
    "sub_category",
    "amount_paid",
]


def openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


def parse_folder_identity(folder_name: str) -> tuple[str, str, str]:
    """Return (deal_id, payment_date_iso, pipeline label)."""
    stem = folder_name.strip()
    pipeline = "LLM"
    if stem.endswith("_sdk"):
        pipeline = "SDK"
        stem = stem[: -len("_sdk")]
    elif stem.endswith("_llm"):
        pipeline = "LLM"
        stem = stem[: -len("_llm")]
    parts = stem.rsplit("_", 1)
    if len(parts) == 2 and len(parts[1]) == 8 and parts[1].isdigit():
        ymd = parts[1]
        iso = f"{ymd[:4]}-{ymd[4:6]}-{ymd[6:8]}"
        return parts[0], iso, pipeline
    return stem, "", pipeline


def _validation_counts(extraction_dir: Path) -> tuple[str, int, int]:
    """Return (status, error_count, warning_count)."""
    report = extraction_dir / "validation_report.md"
    if not report.is_file():
        return "not_run", 0, 0
    text = report.read_text(encoding="utf-8", errors="replace")
    status = "unknown"
    if re.search(r"\*\*STATUS:\s*PASS\*\*", text, re.I) or re.search(
        r"Batch STATUS:\s*PASS", text, re.I
    ):
        status = "pass"
    elif re.search(r"\*\*STATUS:\s*FAIL\*\*", text, re.I) or re.search(
        r"Batch STATUS:\s*FAIL", text, re.I
    ):
        status = "fail"
    elif re.search(r"\bFAIL\b", text[:2000], re.I):
        status = "fail"
    elif re.search(r"\bPASS\b", text[:2000], re.I):
        status = "pass"
    err = len(re.findall(r"\|\s*\*\*ERROR\*\*\s*\|", text, re.I))
    if err == 0:
        err = len(re.findall(r"^\*\*ERROR\*\*", text, re.M | re.I))
    warn = len(re.findall(r"\|\s*\*\*WARN\*\*\s*\|", text, re.I))
    if warn == 0:
        warn = len(re.findall(r"^\*\*WARN\*\*", text, re.M | re.I))
    return status, err, warn


def _kv_maps_from_01(text: str) -> tuple[dict[str, str], dict[str, str]]:
    ident: dict[str, str] = {}
    key_dates: dict[str, str] = {}
    for d in _parse_tables_kv(parse_md_tables(text)):
        if any("deal / trust" in k.lower() for k in d):
            ident = d
        if any(k.strip() == "Payment date" for k in d):
            key_dates = d
    return ident, key_dates


def _listing_cusips_by_class(tables: list[list[list[str]]]) -> dict[str, list[str]]:
    """Economic class → CUSIP list from ``### Tranche by listing``."""
    lt = find_listing_table(tables)
    out: dict[str, list[str]] = {}
    if not lt or len(lt) < 2:
        return out
    hdr = lt[0]
    idx = {_norm_header(c): i for i, c in enumerate(hdr)}

    def gi(row: list[str], *names: str) -> str:
        for n in names:
            i = idx.get(_norm_header(n))
            if i is not None and i < len(row):
                return row[i].strip()
        return ""

    for row in lt[1:]:
        if not any(x.strip() for x in row):
            continue
        ec = gi(row, "Economic class")
        cusip = gi(row, "CUSIP")
        if not ec:
            continue
        if cusip.upper() == "N/A":
            cusip = ""
        if cusip and cusip not in out.setdefault(ec, []):
            out[ec].append(cusip)
        elif not cusip and ec not in out:
            out[ec] = []
    return out


def _pick_interest_payment(row: list[str], cols: dict[str, int | None]) -> str:
    paid = cols.get("interest_payment")
    payable = cols.get("interest_payable")
    div = cols.get("dividend")

    def cell(ix: int | None) -> str:
        if ix is None or ix >= len(row):
            return ""
        return row[ix].strip()

    v = cell(paid)
    if v and v.upper() != "N/A":
        return v
    v = cell(payable)
    if v and v.upper() != "N/A":
        return v
    return cell(div)


def extract_deal_rows(extraction_dir: Path) -> dict[str, Any]:
    """Parse one deal folder into tabular rows for Excel sheets."""
    folder = extraction_dir.name
    deal_id, pay_iso, pipeline = parse_folder_identity(folder)
    notes: list[str] = []

    p01 = extraction_dir / "01_report_metadata.md"
    p02 = extraction_dir / "02_tranche_class_balances.md"
    p03 = extraction_dir / "03_interest_principal_waterfall.md"
    p04 = extraction_dir / "04_extraction_summary.md"
    p05 = extraction_dir / "05_valuation_relevant_fees.md"

    has = {
        "01": p01.is_file(),
        "02": p02.is_file(),
        "03": p03.is_file(),
        "04": p04.is_file(),
        "05": p05.is_file(),
    }
    if not has["01"] or not has["02"]:
        notes.append("missing required 01 or 02")

    ident: dict[str, str] = {}
    key_dates: dict[str, str] = {}
    deal_name = ""
    currency = ""
    payment_01 = ""
    if has["01"]:
        ident, key_dates = _kv_maps_from_01(p01.read_text(encoding="utf-8", errors="replace"))
        for d in (ident, key_dates):
            for k, v in d.items():
                if "deal / trust" in k.lower() or "series name" in k.lower():
                    deal_name = v
                if k.strip().lower() == "currency":
                    currency = v
        payment_01 = key_dates.get("Payment date", "") or key_dates.get("Distribution date", "")

    val_status, val_err, val_warn = _validation_counts(extraction_dir)

    class_rows: list[list[Any]] = []
    if has["02"]:
        t02 = parse_md_tables(p02.read_text(encoding="utf-8", errors="replace"))
        cusips_by_class = _listing_cusips_by_class(t02)
        cb = find_class_balance_table(t02)
        if cb and len(cb) > 1:
            hdr = cb[0]
            iclass = _col_exact(hdr, "Class") or 0
            cols = {
                "beginning_balance": _col_exact(hdr, "Beginning balance"),
                "interest_payment": _col_exact(hdr, "Interest payment"),
                "interest_payable": _col_exact(hdr, "Interest payable"),
                "principal_payment": _col_exact(hdr, "Principal payment"),
                "deferred_interest": _col_exact(hdr, "Deferred interest"),
                "dividend": _col_exact(hdr, "Dividend"),
                "ending_balance": _col_exact(hdr, "Ending balance"),
            }

            def cell(row: list[str], ix: int | None) -> str:
                if ix is None or ix >= len(row):
                    return ""
                return row[ix].strip()

            for row in cb[1:]:
                if len(row) <= iclass:
                    continue
                cname = row[iclass].strip()
                if not cname or cname.lower().startswith("total"):
                    continue
                base = [
                    deal_id,
                    pay_iso,
                    cname,
                    "",
                    _pick_interest_payment(row, cols),
                    cell(row, cols["principal_payment"]),
                    cell(row, cols["beginning_balance"]),
                    cell(row, cols["deferred_interest"]),
                    cell(row, cols["ending_balance"]),
                ]
                cusips = cusips_by_class.get(cname) or []
                if cusips:
                    for cusip in cusips:
                        r = list(base)
                        r[3] = cusip
                        class_rows.append(r)
                else:
                    class_rows.append(base)
        else:
            notes.append("no primary class table in 02")

    fee_rows: list[list[Any]] = []
    fee_text = ""
    if has["05"]:
        fee_text = p05.read_text(encoding="utf-8", errors="replace")
    elif has["03"]:
        fee_text = p03.read_text(encoding="utf-8", errors="replace")
        notes.append("fees from 03 (05 missing)")
    vf = find_valuation_fees_table(parse_md_tables(fee_text)) if fee_text else None
    if vf and len(vf) > 1:
        header = vf[0]
        tc = fee_type_col(header)
        mc = fee_main_col(header)
        ac = fee_amount_col(header)
        for row in fee_data_rows(vf):
            fee_rows.append(
                [
                    deal_id,
                    pay_iso,
                    row[mc].strip() if mc is not None and mc < len(row) else "",
                    row[tc].strip() if tc < len(row) else "",
                    row[ac].strip() if ac < len(row) else "",
                ]
            )
    elif fee_text:
        notes.append("no valuation fees table parsed")

    summary_row = [
        folder,
        deal_id,
        pay_iso,
        pipeline,
        "Y" if has["01"] else "N",
        "Y" if has["02"] else "N",
        "Y" if has["03"] else "N",
        "Y" if has["04"] else "N",
        "Y" if has["05"] else "N",
        val_status,
        val_err,
        val_warn,
        deal_name,
        payment_01,
        currency,
        len(class_rows),
        len(fee_rows),
        "; ".join(notes),
    ]

    return {
        "summary": summary_row,
        "classes": class_rows,
        "fees": fee_rows,
    }


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in rows[:500]:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1] or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def build_workbook_bytes(
    deal_dirs: list[Path],
    *,
    tranche_mapper: TrancheMapper | None = None,
) -> bytes:
    """Build multi-sheet xlsx for the given extraction directories."""
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    summaries: list[list[Any]] = []
    classes: list[list[Any]] = []
    fees: list[list[Any]] = []

    for d in sorted(deal_dirs, key=lambda p: p.name.lower()):
        if not d.is_dir():
            continue
        try:
            parsed = extract_deal_rows(d.resolve())
        except OSError as e:
            deal_id, pay_iso, pipeline = parse_folder_identity(d.name)
            summaries.append(
                [
                    d.name,
                    deal_id,
                    pay_iso,
                    pipeline,
                    "N",
                    "N",
                    "N",
                    "N",
                    "N",
                    "error",
                    0,
                    0,
                    "",
                    "",
                    "",
                    0,
                    0,
                    f"read error: {e}",
                ]
            )
            continue
        summaries.append(parsed["summary"])
        classes.extend(parsed["classes"])
        fees.extend(parsed["fees"])

    class_headers = list(_CLASS_EXPORT_HEADERS)
    if tranche_mapper is not None:
        classes = [tranche_mapper.enrich_class_row(r) for r in classes]
        class_headers = class_headers + list(CLASS_MAP_EXTRA_HEADERS)

    ws_cls = wb.create_sheet("Classes", 0)
    _write_sheet(ws_cls, class_headers, classes)
    ws_fee = wb.create_sheet("Valuation fees")
    _write_sheet(ws_fee, _VALUATION_FEE_HEADERS, fees)
    ws_sum = wb.create_sheet("Summary")
    _write_sheet(ws_sum, _SUMMARY_HEADERS, summaries)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def discover_deal_dirs(
    output_root: Path,
    *,
    source: str = "llm",
    folder_names: list[str] | None = None,
    max_deals: int = 0,
) -> list[Path]:
    """Resolve folder paths under output_root (same rules as batch validate)."""
    root = output_root.resolve()
    if folder_names:
        out: list[Path] = []
        for name in folder_names:
            n = name.strip()
            if not n:
                continue
            p = (root / n).resolve()
            if p.is_dir() and _DEAL_FOLDER.match(p.name):
                out.append(p)
        return sorted(out, key=lambda x: x.name.lower())[: max_deals or None]

    src = (source or "llm").strip().lower()
    found: list[Path] = []
    if not root.is_dir():
        return found
    for p in root.iterdir():
        if not p.is_dir() or not _DEAL_FOLDER.match(p.name):
            continue
        if src == "sdk" and not p.name.endswith("_sdk"):
            continue
        if src == "llm" and p.name.endswith("_sdk"):
            continue
        if (p / "01_report_metadata.md").is_file() or (p / "02_tranche_class_balances.md").is_file():
            found.append(p)
    found.sort(key=lambda x: x.name.lower())
    if max_deals and max_deals > 0:
        return found[:max_deals]
    return found


def default_export_filename(deal_count: int, source: str = "llm") -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return f"noteval_export_{source}_{deal_count}deals_{ts}.xlsx"


def main() -> int:
    ap = argparse.ArgumentParser(description="Export noteval markdown folders to one Excel workbook.")
    ap.add_argument(
        "extraction_dirs",
        nargs="+",
        type=Path,
        help="One or more deal output folders (01/02 required for full rows)",
    )
    ap.add_argument("-o", "--output", type=Path, required=True, help="Output .xlsx path")
    ap.add_argument(
        "--map-tranches",
        action="store_true",
        help="Add moodystrancheid columns via map_tranches (DB or cache)",
    )
    ap.add_argument(
        "--tranche-cache",
        type=Path,
        help="JSON cache from map_tranches.py --prefetch (optional with --map-tranches)",
    )
    ap.add_argument(
        "--no-tranche-db",
        action="store_true",
        help="With --map-tranches, use only --tranche-cache (no SQL)",
    )
    args = ap.parse_args()
    dirs = [d.resolve() for d in args.extraction_dirs]
    mapper: TrancheMapper | None = None
    if args.map_tranches:
        from noteval_01_xml import ensure_01_xml_for_mapping

        for d in dirs:
            try:
                p01 = ensure_01_xml_for_mapping(d)
                print(f"01 XML: {p01}", file=sys.stderr)
            except FileNotFoundError as e:
                print(f"warn: {e}", file=sys.stderr)
        mapper = TrancheMapper(
            cache_file=args.tranche_cache.resolve() if args.tranche_cache else None,
            use_db=not args.no_tranche_db,
        )
    data = build_workbook_bytes(dirs, tranche_mapper=mapper)
    out = args.output.resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(data)
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
